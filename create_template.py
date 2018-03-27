# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

from openpyxl import Workbook


def create_template():
    templ=input('Number of required templates: ')
    #managing wrong inputs
    try:
        int(templ)
    except ValueError:
        print 'Not a number, start again'
    for q in range(templ):#iteration to create differenet templates
        num=q+1
        name=raw_input('Name of the file '+str(num)+': ')
        filename=name+'.xlsx'
        '''print 'For ',filename,':'
        years=input('\t How many years of data: ')
        try:
            int(years)
        except ValueError:
            print 'Not a number, start again'
            break
        initial=input('\t From year(incl.): ')
        try:
            int(initial)
        except ValueError:
            print 'Not a number, start again'
            break
        final=input('\t Until year(incl.): ')
        try:
            int(final)
        except ValueError:
            print 'Not a number, start again'
            break'''
        numheight=input('\t How many different heights: ')
        try:
            int(numheight)
        except ValueError:
            print 'Not a number'
        if numheight>5:
            print 'Maximum 5 different heights, start again'
            break
        else:
            '''yearset=[]#list of all years of data which will become a worksheet
            for x in range(final-initial+1):
                yearset.append(initial+x)'''
            heights=[]
            for j in range(numheight):#iteration to introduce height of instruments
                h1=raw_input('\t \t Height (meters)'+str(j+1)+': ')
                #managing wrong inputs
                try:
                    float(h1)
                except ValueError:
                    print 'Not a number, start again'
                    break
                heights.append(h1)
            date=raw_input('\t Do you use date&time stamp separately?(y/n): ')
            '''d=[]
            for k in range(years):#iteration to create worksheets with headings
                d.append('ws'+str(k)+'')'''
            if date=='y': 
                wb=Workbook()
                #colums=5+(2*numheight)
                b=-1#a it's to know position in yearset
                '''for item in d:#iteration to create worksheets with headings'''
                b+=1
                item=wb.create_sheet()
                item.title='Data'
                item['A1']='DAY'
                item['B1']='TIME'
                item['C1']='TEMPERATURE(C)'
                item['D1']='PRESSURE(hPa)'
                item['E1']='HUMIDITY(%)'
                item['F1']='SIGNIFICANT_WAVE(m)'
                item['G1']='WAVELENGHT(m)'
                item['H1']='TSEA(C)'
                for i in range(numheight):
                    a=-1+(2*i)
                    a+=1
                    item.cell(row=0, column=(8+a)).value='WIND SPEED'+str(heights[i])+'(m/s)'
                    item.cell(row=0, column=(8+1+a)).value='WIND DIRECTION'+str(heights[i])+'(degrees)'
                ws=wb.get_sheet_by_name('Sheet')
                wb.remove_sheet(ws)
                wb.save(filename=filename)
                
            elif date=='n':
                wb=Workbook()
                #colums=4+(2*numheight)
                b=-1#a it's to know position in yearset
                '''for item in d:'''
                b+=1
                item=wb.create_sheet()
                item.title='Data'
                item['A1']='DATE&TIME STAMP'
                item['B1']='TEMPERATURE(C)'
                item['C1']='PRESSURE(hPa)'
                item['D1']='HUMIDITY(%)'
                item['E1']='SIGNIFICANT_WAVE(m)'
                item['F1']='WAVELENGHT(m)'
                item['G1']='TSEA(C)'
                for i in range(numheight):
                    a=-1+(2*i)
                    a+=1
                    item.cell(row=0, column=(7+a)).value='WIND SPEED'+str(heights[i])+'(m/s)'
                    item.cell(row=0, column=(7+1+a)).value='WIND DIRECTION'+str(heights[i])+'(degrees)'
                ws=wb.get_sheet_by_name('Sheet')
                wb.remove_sheet(ws)
                wb.save(filename=filename)
            print '--------------'+filename+' ready-------------------'


